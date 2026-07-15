"""Groww portfolio import.

Two optional ways for a user to bring their Groww holdings into AlphaHive — the
portfolio is never required; the app works as a normal research chat without it.

1. Official Groww Trade API (`GET /v1/holdings/user`): the user generates a daily
   access token from Groww's Trading APIs page and pastes it in. We map the
   response straight onto our holding shape. Requires a Groww API subscription.
   Docs: https://groww.in/trade-api/docs/curl/portfolio

2. Statement upload: the user exports Holdings / P&L as CSV or Excel from Groww's
   Reports section and uploads the file. We parse it heuristically (broker export
   column names vary), so no subscription is needed.

Every parser returns the same normalized shape:
    {"symbol": str, "shares": float, "average_buy_price": float}
"""

import csv
import io
import logging
from typing import Any, Dict, List

import httpx

logger = logging.getLogger("groww-service")

GROWW_HOLDINGS_URL = "https://api.groww.in/v1/holdings/user"

# Header aliases we accept in an uploaded CSV/XLSX, lower-cased. Groww's own
# exports and common community exports use varying names for the same column.
_SYMBOL_KEYS = {"symbol", "trading_symbol", "tradingsymbol", "stock", "stock name", "scrip", "instrument", "company"}
_QTY_KEYS = {"quantity", "qty", "shares", "net quantity", "holding quantity", "units"}
_AVG_KEYS = {"average_price", "average price", "avg price", "avg. price", "avg cost", "buy average", "average buy price", "avg buy price"}


class GrowwImportError(Exception):
    """Raised when a Groww import cannot be completed (bad token, unparseable file)."""


class GrowwService:
    async def fetch_holdings_via_api(self, access_token: str) -> List[Dict[str, Any]]:
        """Fetch live holdings from the official Groww Trade API using a user's
        access token. Returns normalized holdings. Raises GrowwImportError on
        auth/HTTP failure."""
        token = (access_token or "").strip()
        if not token:
            raise GrowwImportError("A Groww access token is required.")

        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {token}",
            "X-API-VERSION": "1.0",
        }
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(GROWW_HOLDINGS_URL, headers=headers)
        except Exception as e:  # noqa: BLE001
            raise GrowwImportError(f"Could not reach Groww API: {e}") from e

        if resp.status_code in (401, 403):
            raise GrowwImportError(
                "Groww rejected the access token (expired or invalid). Groww tokens "
                "expire daily at 6 AM — generate a fresh one and try again."
            )
        if resp.status_code != 200:
            raise GrowwImportError(f"Groww API returned HTTP {resp.status_code}.")

        try:
            payload = resp.json()
        except Exception as e:  # noqa: BLE001
            raise GrowwImportError(f"Groww API returned invalid JSON: {e}") from e

        # The holdings list may be under a top-level key (e.g. "holdings"/"payload")
        # or be the payload itself.
        rows = self._extract_rows(payload)
        holdings = []
        for row in rows:
            symbol = row.get("trading_symbol") or row.get("symbol")
            qty = row.get("quantity")
            avg = row.get("average_price")
            if not symbol or qty in (None, 0):
                continue
            holdings.append(
                {
                    "symbol": str(symbol).upper().strip(),
                    "shares": float(qty),
                    "average_buy_price": float(avg or 0.0),
                }
            )
        if not holdings:
            raise GrowwImportError("Groww returned no stock holdings for this account.")
        return holdings

    @staticmethod
    def _extract_rows(payload: Any) -> List[Dict[str, Any]]:
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            for key in ("holdings", "payload", "data", "result"):
                val = payload.get(key)
                if isinstance(val, list):
                    return val
                if isinstance(val, dict):
                    inner = val.get("holdings")
                    if isinstance(inner, list):
                        return inner
        return []

    def parse_holdings_file(self, filename: str, content: bytes) -> List[Dict[str, Any]]:
        """Parse an uploaded Groww holdings export (CSV or XLSX) into normalized
        holdings. Raises GrowwImportError if no usable rows are found."""
        name = (filename or "").lower()
        if name.endswith(".xlsx") or name.endswith(".xls"):
            rows = self._read_excel_rows(content)
        else:
            rows = self._read_csv_rows(content)

        holdings = self._rows_to_holdings(rows)
        if not holdings:
            raise GrowwImportError(
                "Couldn't find holdings in that file. Expected columns like "
                "Symbol/Trading Symbol, Quantity, and Average Price."
            )
        return holdings

    @staticmethod
    def _read_csv_rows(content: bytes) -> List[Dict[str, str]]:
        text = content.decode("utf-8-sig", errors="replace")
        # Groww exports sometimes have preamble lines before the header row; find the
        # line that looks like a header (contains a symbol-ish and a quantity-ish token).
        lines = text.splitlines()
        header_idx = 0
        for i, line in enumerate(lines[:25]):
            low = line.lower()
            if any(k in low for k in _SYMBOL_KEYS) and any(k in low for k in _QTY_KEYS):
                header_idx = i
                break
        reader = csv.DictReader(lines[header_idx:])
        return [dict(r) for r in reader]

    @staticmethod
    def _read_excel_rows(content: bytes) -> List[Dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover
            raise GrowwImportError("Excel support is not installed on the server.") from e

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        grid = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
        if not grid:
            return []
        # Locate the header row (first row containing a symbol-ish + quantity-ish cell).
        header_idx = 0
        for i, row in enumerate(grid[:25]):
            low = [str(c).lower() for c in row]
            if any(any(k in cell for k in _SYMBOL_KEYS) for cell in low) and any(
                any(k in cell for k in _QTY_KEYS) for cell in low
            ):
                header_idx = i
                break
        headers = [str(c).strip() for c in grid[header_idx]]
        rows = []
        for row in grid[header_idx + 1:]:
            rows.append({headers[j]: row[j] for j in range(min(len(headers), len(row)))})
        return rows

    @classmethod
    def _rows_to_holdings(cls, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        holdings = []
        for row in rows:
            norm = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
            symbol = cls._pick(norm, _SYMBOL_KEYS)
            qty = cls._pick(norm, _QTY_KEYS)
            avg = cls._pick(norm, _AVG_KEYS)
            if not symbol:
                continue
            shares = cls._to_float(qty)
            if shares <= 0:
                continue
            holdings.append(
                {
                    "symbol": str(symbol).upper().strip(),
                    "shares": shares,
                    "average_buy_price": cls._to_float(avg),
                }
            )
        return holdings

    @staticmethod
    def _pick(norm: Dict[str, Any], keys: set) -> Any:
        for k in keys:
            if k in norm and norm[k] not in (None, ""):
                return norm[k]
        return None

    @staticmethod
    def _to_float(val: Any) -> float:
        if val is None:
            return 0.0
        try:
            return float(str(val).replace(",", "").replace("₹", "").strip() or 0.0)
        except (TypeError, ValueError):
            return 0.0


groww_service = GrowwService()
